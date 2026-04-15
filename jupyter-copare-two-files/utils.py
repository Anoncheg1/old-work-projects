# pip install openpyxl xlrd notmuch
import os
import re
import csv
import openpyxl
import xlrd
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import InvalidFileException
from os.path import basename
from openpyxl import Workbook
from datetime import date, timedelta
import notmuch
import logging
import traceback
from typing import Generator, Optional
# import pyclamd

# def xlsx_scan_for_viruses(file_path):
#     clamd = pyclamd.ClamdNetworkSocket()
#     result = clamd.scan(file_path)
#     if result:
#         raise ValueError("XLSX file contains malware")

logger = logging.root


def csv_row_iterator(file_path: str) -> list:
    with open(file_path, 'r', newline='', encoding='utf-8', errors='ignore') as fhander:
        reader = csv.reader(fhander)
        for row in reader:
            yield row


def xlsx_row_iterator(file_path: str, sheet: str=None) -> list:
    """Can't detect max_row properly use:
    all(e is None for e in row)"""
    # - Read the XLSX file
    try:
        wb: Workbook = openpyxl.load_workbook(file_path, read_only=True,
                                              keep_vba=False, data_only=True,
                                              keep_links=False)  # , rich_text=False
        # - check XLSX for bombs and viruses:
        xlsx_check_for_macros(wb)
        # - Get the active worksheet
        if sheet:
            sheet: Worksheet = wb[sheet]
        else:
            sheet: Worksheet = wb.active  # wb['Sheet1']
        for row in sheet.iter_rows(min_row=1, max_row=None, min_col=None,
                                   max_col=None, values_only=True):
            yield row
    except InvalidFileException as e:  # try to open as xls
        wb = xlrd.open_workbook(file_path)
        # Access sheet by index or name
        sheet = wb.sheet_by_index(0)  # Access first sheet
        # Iterate through rows
        for row_index in range(sheet.nrows):
            yield sheet.row_values(row_index)


def create_row_iterator(file_path: str) -> list:
    """CSV/XLSX"""
    # - Read file
    if str.lower(file_path).endswith('.csv'):
        return csv_row_iterator(file_path)
    elif str.lower(file_path).endswith('.xlsx'):
        return xlsx_row_iterator(file_path)
    else:
        raise ValueError("File is not CSV/XLSX")

def xlsx_check_for_macros(wb: Workbook):
    if wb.vba_archive or wb._external_links:
        raise ValueError("XLSX file contains macros or links")


def xlsx_write(header: list[str] | None, data: list[list], filename: str = 'output.xlsx'):
    """
    Write an XLSX file with the given header and data.

    Args:
        header (list[str]): The header row.
        data (list[list]): The data rows.
        filename (str, optional): The filename to write to. Defaults to 'output.xlsx'.
    """
    if filename is None:
        raise ValueError("Filename should be specified.")
    # Create a new workbook
    wb: Workbook = Workbook()

    # Get the active worksheet
    ws = wb.active

    # Write the header row
    if header:
        for i, value in enumerate(header):
            ws.cell(row=1, column=i+1, value=value)

    # Write the data rows
    for i, row in enumerate(data, start=2):
        for j, value in enumerate(row):
            ws.cell(row=i, column=j+1, value=value)

    # Save the workbook to a file
    wb.save(filename)


def write_csv_file(filename: str, data: list[list], header:list[str]=None):
    """
    Write a CSV file with the given data.

    Args:
        filename (str): The name of the CSV file to write.
        data (list of lists): The data to write to the CSV file.
        header (list of strings): first line of column name to write.

    Returns:
        None
    """
    with open(filename, 'w', newline='', encoding='utf-8', errors='ignore') as csvfile:
        writer = csv.writer(csvfile)
        if header:
            writer.writerow(header)
        writer.writerows(data)


def date_extract(file_path: str) -> str:
    """
    Returns:         str: The extracted date in the format m/d/year.
    """
    pattern1 = r'(\d{4})_(\d{2})_(\d{2})'
    pattern2 = r'\b(\d{2}-[A-Za-z]{3}-\d{2})\b'
    pattern3 = r'(\d{4})-(\d{2})-(\d{2})'
    fname = basename(file_path)
    match1 = re.search(pattern1, fname)
    match2 = re.search(pattern2, fname)
    match3 = re.search(pattern3, fname)
    if match1:
        year, month, day = match1.groups()
        return f"{month}/{day}/{year}"
    elif match2:
        day, month_abbr, year = match2.group(1).split('-')
        day = day
        year = '20' + year  # Assuming '25' means 2025
        # Dictionary to convert month abbreviations to numbers
        month_dict = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        # Convert month abbreviation to number
        month = month_dict[month_abbr.lower()]
        return f"{month}/{day}/{year}"
    elif match3:
        year, month, day = match3.groups()
        return f"{month}/{day}/{year}"
    else:
        raise ValueError(f"Not able to exract date from file name: {fname}")


def date_to_intlist(date_str: str):
    """ 01/06/2025 -> [ 1, 6, 2025 ]"""
    date_list: list = date_str.split('/')
    return int(date_list[0]), int(date_list[1]), int(date_list[2])


def date_add_day(month, day, year):
    dt = date(year, month, day) + timedelta(days=1)
    return dt.month, dt.day, dt.year


def date_add_months(day, month, year, months=12):
    dt = date(year, month, day) + timedelta(months=months)
    return dt.month, dt.day, dt.year


def date_is_within_n_months(current_month, current_year, check_month, check_year, nm=12):
    """
    Returns True if the check date is not more than 12 months in the future from the current date.

    Args:
        current_month (int): Current month (1-12)
        current_year (int): Current year
        check_month (int): Month to check (1-12)
        check_year (int): Year to check

    Returns:
        bool: True if check date is within 12 months, False otherwise
    """
    current_date = current_year * nm + current_month
    check_date = check_year * nm + check_month
    return check_date >= current_date and check_date - current_date <= 12

# def date_not_later():


def header_to_headerdict(header_values, columns_to_process):
    """ Get indexes for columns (case ignored)."""
    # ^ ['TRADE DATE', 'HUB', 'PRODUCT', 'STRIP', 'CONTRACT', 'CONTRACT TYPE', ...]
    header_values: list = [s.casefold() for s in header_values] # ignore case
    header_dict: dict = {}
    for col in columns_to_process:
        col_cf = col.casefold()
        if col_cf in header_values:
            header_dict[col] = header_values.index(col_cf)
        else:
            raise ValueError(f"Incorrect table, column {col} not found.")
    return header_dict
# Alternative: without checking if earlier: (check_year - current_year) * 12 + check_month - current_month <= 12:
    # current_date = current_year * 12 + current_month
    # check_date = check_year * 12 + check_month
    # return abs(check_date - current_date) <= 12 and (check_date >= current_date or check_date + 12 >= current_date)


def tag_email_processed(msg: notmuch.Message, db: notmuch.Database, tag: str="tosftp"):
    "Notmuch"
    try:
        res = db.begin_atomic()
        if res != notmuch.errors.STATUS.SUCCESS:
            raise ValueError(f'Atomic notmuch is failed.')

        msg.remove_tag(tag)
        msg.add_tag("processed")
        res = msg.tags_to_maildir_flags()
        if res != notmuch.errors.STATUS.SUCCESS:
            raise ValueError(f'Update notmuch tags failed.')

        res = db.end_atomic()
        if res != notmuch.errors.STATUS.SUCCESS:
            raise ValueError(f'Atomic notmuch is failed.')
    except Exception as e:
        # logger.error(f'Message not send: {msg}', exc_info=True)
        logger.critical('An error occurred2: %s: %s\n%s' % (e.__class__,
                                                        e,
                                                        traceback.format_exc()), stack_info=True)
        db.close()
        sys.exit(1)




def list_files_in_directory(directory_path: str) -> Optional[Generator[str, None, None]]:
    """
    List all files in the given directory and yield their full paths.

    Parameters:
    - directory_path: str, the path to the directory to be searched

    Returns:
    Iterator[str] or None: An iterator of full file paths, or None if the directory doesn't exist
    """
    # Ensure the provided path is absolute
    abs_directory_path: str = os.path.abspath(directory_path)

    # Check if the directory exists
    if not os.path.isdir(abs_directory_path):
        print(f"Error: The directory '{abs_directory_path}' does not exist.")
        return None

    # Iterate over all entries in the directory
    for entry in os.listdir(abs_directory_path):
        full_path: str = os.path.join(abs_directory_path, entry)

        # Check if the entry is a file (not a directory)
        if os.path.isfile(full_path):
            yield full_path


def parse_emails(emails_string: str) -> list:
    "for distrubution table of clients"
    def is_valid_email(email):
        parts = email.split('@')
        return len(parts) == 2 and all(parts) and '.' in parts[1]

    emails = emails_string.replace('(', '').replace(')', '').replace(';', ' ')
    emails = emails.split()
    return [email for email in emails if is_valid_email(email)]
